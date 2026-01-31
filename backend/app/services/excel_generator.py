
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from app.models import SeatingResult, HallSeating
from collections import defaultdict

# Shared styles
TITLE_FONT = Font(name='Times New Roman', size=12, bold=True)
HEADER_FONT = Font(name='Times New Roman', size=10, bold=True)
DATA_FONT = Font(name='Times New Roman', size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
WRAP_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER_STYLE = Side(style='thin')
FULL_BORDER = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)


def get_snake_seat_number(row_idx: int, col_idx: int, num_rows: int) -> int:
    """Calculate seat number using vertical snake pattern."""
    if col_idx % 2 == 0:  # Even columns (0, 2, 4...): top to bottom
        return (col_idx * num_rows) + row_idx + 1
    else:  # Odd columns (1, 3, 5...): bottom to top
        return (col_idx * num_rows) + (num_rows - row_idx)


def get_exam_info(seating_result: SeatingResult) -> tuple:
    """Extract exam date and session from first student."""
    exam_date = "NOV/DEC 2025"
    session = "FN"
    date_str = ""
    for hall_seating in seating_result.halls:
        for row in hall_seating.grid:
            for seat in row:
                if seat.student:
                    exam_date = seat.student.examDate or exam_date
                    session = seat.student.session or session
                    date_str = exam_date
                    return exam_date, session, date_str
    return exam_date, session, date_str


def generate_hall_wise_excel(seating_result: SeatingResult) -> BytesIO:
    """
    Generate Excel file with all sheets matching the reference format.
    Sheets: SEATING, HALL ALLO, NB, aud seating, Sheet1
    """
    output = BytesIO()
    wb = Workbook()
    
    exam_date, session, date_str = get_exam_info(seating_result)
    
    # Collect data for all sheets
    hall_data = []  # For HALL ALLO sheet
    dept_data = defaultdict(lambda: {'students': [], 'halls': defaultdict(list)})  # For NB sheet
    regular_halls = []
    auditorium_halls = []
    
    for hall_seating in seating_result.halls:
        if hall_seating.studentsCount == 0:
            continue
        
        hall = hall_seating.hall
        is_auditorium = 'AUD' in hall.name.upper() or hall.name.upper().startswith('A')
        
        if is_auditorium:
            auditorium_halls.append(hall_seating)
        else:
            regular_halls.append(hall_seating)
        
        # Collect data for HALL ALLO and NB sheets
        subject_counts = defaultdict(int)
        for row in hall_seating.grid:
            for seat in row:
                if seat.student:
                    subject_counts[seat.student.subjectCode] += 1
                    dept = seat.student.department
                    dept_data[dept]['students'].append(seat.student.registerNumber)
                    dept_data[dept]['halls'][hall.name].append(seat.student.registerNumber)
        
        hall_data.append({
            'hall': hall.name,
            'subjects': subject_counts,
            'total': hall_seating.studentsCount
        })
    
    # === SHEET 1: SEATING (Regular halls) ===
    ws_seating = wb.active
    ws_seating.title = "SEATING"
    _write_seating_sheet(ws_seating, regular_halls, exam_date, session)
    
    # === SHEET 2: HALL ALLO ===
    ws_hall_allo = wb.create_sheet("HALL ALLO")
    _write_hall_allo_sheet(ws_hall_allo, hall_data, date_str, session)
    
    # === SHEET 3: NB (Department breakdown) ===
    ws_nb = wb.create_sheet("NB")
    _write_nb_sheet(ws_nb, dept_data, date_str, session)
    
    # === SHEET 4: aud seating (Auditorium layout) ===
    if auditorium_halls:
        ws_aud = wb.create_sheet("aud seating")
        _write_auditorium_sheet(ws_aud, auditorium_halls, exam_date, session)
    
    wb.save(output)
    output.seek(0)
    return output


def _write_seating_sheet(ws, halls, exam_date, session):
    """Write the SEATING sheet with hall sketches."""
    current_row = 1
    row_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
    
    # Configure A4 page settings and margins
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 means auto
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    
    hall_start_row = current_row
    
    for idx, hall_seating in enumerate(halls):
        hall = hall_seating.hall
        grid = hall_seating.grid
        num_cols = hall.columns
        num_rows = hall.rows
        excel_data_cols = num_cols * 2
        
        # Title row
        title_text = "GCE : : ERODE - 638 316 - ANNA UNIVERSITY EXAMS - HALL  SKETCH - NOV/DEC 2025"
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=excel_data_cols)
        title_cell = ws.cell(row=current_row, column=1, value=title_text)
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN
        current_row += 1
        
        # Hall info row
        ws.cell(row=current_row, column=1, value=f"HALL NO : {hall.name}").font = HEADER_FONT
        date_col = (excel_data_cols // 2) + 1
        ws.cell(row=current_row, column=date_col, value=f"Date & Session :  {exam_date} {session}").font = HEADER_FONT
        current_row += 1
        
        # Column headers
        for col_idx in range(num_cols):
            excel_col = (col_idx * 2) + 1
            roman = row_numerals[col_idx] if col_idx < len(row_numerals) else str(col_idx + 1)
            header_text = f"{roman} - ROW with Seat No"
            cell = ws.cell(row=current_row, column=excel_col, value=header_text)
            cell.font = HEADER_FONT
            cell.alignment = WRAP_CENTER_ALIGN
            ws.column_dimensions[get_column_letter(excel_col)].width = 15
            ws.column_dimensions[get_column_letter(excel_col + 1)].width = 6
        current_row += 1
        
        # Data rows and department tracking
        dept_students = defaultdict(list)
        
        for row_idx in range(num_rows):
            for col_idx in range(num_cols):
                excel_col = (col_idx * 2) + 1
                data_row = current_row + row_idx
                
                seat = grid[row_idx][col_idx] if row_idx < len(grid) and col_idx < len(grid[row_idx]) else None
                
                reg_cell = ws.cell(row=data_row, column=excel_col)
                seat_num_cell = ws.cell(row=data_row, column=excel_col + 1)
                seat_number = get_snake_seat_number(row_idx, col_idx, num_rows)
                
                if seat and seat.student:
                    reg_cell.value = seat.student.registerNumber
                    dept_students[seat.student.department].append(seat.student.registerNumber)
                else:
                    reg_cell.value = ""
                
                seat_num_cell.value = seat_number
                
                for c in [reg_cell, seat_num_cell]:
                    c.font = DATA_FONT
                    c.alignment = CENTER_ALIGN
                    c.border = FULL_BORDER
        
        current_row += num_rows + 1
        
        # Department summary
        for dept, students in sorted(dept_students.items()):
            students_sorted = sorted(students)
            range_text = students_sorted[0] if len(students_sorted) == 1 else f"{students_sorted[0]} TO {students_sorted[-1]}"
            
            ws.cell(row=current_row, column=1, value=dept).font = DATA_FONT
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            range_cell = ws.cell(row=current_row, column=2, value=range_text)
            range_cell.font = DATA_FONT
            range_cell.alignment = LEFT_ALIGN
            ws.cell(row=current_row, column=excel_data_cols, value=len(students)).font = DATA_FONT
            current_row += 1
        
        # TOTAL row
        ws.cell(row=current_row, column=excel_data_cols - 1, value="TOTAL").font = HEADER_FONT
        ws.cell(row=current_row, column=excel_data_cols, value=hall_seating.studentsCount).font = HEADER_FONT
        current_row += 1
        # Add page break after every 4 halls (except the last set)
        is_page_break = (idx + 1) % 4 == 0 and idx < len(halls) - 1
        is_last_hall = idx == len(halls) - 1
        
        # Add separator line (thick border) between halls - but NOT before page breaks or after last hall
        if not is_page_break and not is_last_hall:
            separator_border = Border(bottom=Side(style='medium'))
            for col in range(1, excel_data_cols + 1):
                ws.cell(row=current_row, column=col).border = separator_border
            current_row += 2  # Leave a blank row after the separator line
        
        if is_page_break:
            ws.row_breaks.append(Break(id=current_row - 1))


def _write_hall_allo_sheet(ws, hall_data, date_str, session):
    """Write the HALL ALLO sheet with hall allocation summary."""
    # Headers
    headers = ["HALL", "DEPT", "SUBJECT", "TOTAL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = FULL_BORDER
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 8
    
    current_row = 2
    for hall_info in hall_data:
        hall_name = hall_info['hall']
        subjects = hall_info['subjects']
        total = hall_info['total']
        
        num_subjects = len(subjects)
        # Reserve rows: subjects + 1 empty row (for spacing before TOTAL row)
        data_rows = max(num_subjects, 4)  # Minimum 4 rows for visual consistency
        
        start_row = current_row
        end_row = current_row + data_rows - 1
        
        # Merge HALL column for all subject rows
        if data_rows > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        
        hall_cell = ws.cell(row=start_row, column=1, value=hall_name)
        hall_cell.font = DATA_FONT
        hall_cell.alignment = CENTER_ALIGN
        hall_cell.border = FULL_BORDER
        
        # Apply borders to merged cells
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=1).border = FULL_BORDER
            ws.cell(row=r, column=2).border = FULL_BORDER
            ws.cell(row=r, column=3).border = FULL_BORDER
            ws.cell(row=r, column=4).border = FULL_BORDER
        
        # Write subjects
        for i, (subject, count) in enumerate(subjects.items()):
            ws.cell(row=start_row + i, column=3, value=subject).font = DATA_FONT
            ws.cell(row=start_row + i, column=4, value=count).font = DATA_FONT
        
        current_row = end_row + 1
        
        # Date/session and TOTAL row
        date_cell = ws.cell(row=current_row, column=1, value=f"{date_str} {session}")
        date_cell.font = DATA_FONT
        date_cell.border = FULL_BORDER
        ws.cell(row=current_row, column=2).border = FULL_BORDER
        ws.cell(row=current_row, column=3, value="TOTAL:").font = HEADER_FONT
        ws.cell(row=current_row, column=3).border = FULL_BORDER
        ws.cell(row=current_row, column=4, value=total).font = HEADER_FONT
        ws.cell(row=current_row, column=4).border = FULL_BORDER
        current_row += 1


def _write_nb_sheet(ws, dept_data, date_str, session):
    """Write the NB sheet with department-wise breakdown."""
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="GCE : : ERODE - ANNA UNIVERSITY EXAMS - NOV/DEC 2025")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN
    
    ws.merge_cells('A2:E2')
    ws.cell(row=2, column=1, value="HALL ALLOCATION").font = HEADER_FONT
    
    ws.cell(row=4, column=1, value=f"DATE & SESSION  :  {date_str} {session}").font = HEADER_FONT
    
    current_row = 6
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    
    for dept, data in sorted(dept_data.items()):
        students = sorted(data['students'])
        total = len(students)
        
        # Department header
        ws.cell(row=current_row, column=1, value=f"DEPARTMENT / SEMESTER  :    {dept}").font = HEADER_FONT
        ws.cell(row=current_row, column=2, value="TOTAL :").font = HEADER_FONT
        ws.cell(row=current_row, column=3, value=total).font = HEADER_FONT
        current_row += 1
        
        # Table header row with borders
        header_reg = ws.cell(row=current_row, column=1, value="REGISTER NUMBERS")
        header_reg.font = HEADER_FONT
        header_reg.border = FULL_BORDER
        header_count = ws.cell(row=current_row, column=2, value="COUNT")
        header_count.font = HEADER_FONT
        header_count.border = FULL_BORDER
        header_hall = ws.cell(row=current_row, column=3, value="HALL")
        header_hall.font = HEADER_FONT
        header_hall.border = FULL_BORDER
        current_row += 1
        
        # Hall-wise register numbers with borders
        for hall_name, hall_students in sorted(data['halls'].items()):
            hall_students_sorted = sorted(hall_students)
            count = len(hall_students_sorted)
            
            if count == 1:
                range_text = hall_students_sorted[0]
            elif count <= 3:
                range_text = ", ".join(hall_students_sorted)
            else:
                range_text = f"{hall_students_sorted[0]} TO {hall_students_sorted[-1]}"
            
            reg_cell = ws.cell(row=current_row, column=1, value=range_text)
            reg_cell.font = DATA_FONT
            reg_cell.border = FULL_BORDER
            count_cell = ws.cell(row=current_row, column=2, value=count)
            count_cell.font = DATA_FONT
            count_cell.border = FULL_BORDER
            hall_cell = ws.cell(row=current_row, column=3, value=hall_name)
            hall_cell.font = DATA_FONT
            hall_cell.border = FULL_BORDER
            current_row += 1
        
        current_row += 1  # Empty row between departments


def _write_auditorium_sheet(ws, auditorium_halls, exam_date, session):
    """Write the auditorium seating sheet (3 cols x 9 rows, 25 seats)."""
    current_row = 1
    row_numerals = ["I", "II", "III"]
    
    for hall_seating in auditorium_halls:
        hall = hall_seating.hall
        grid = hall_seating.grid
        
        # Auditorium specific: 3 columns, 9 rows, 25 total seats
        # The pattern has XXX markers for missing seats
        num_cols = 3
        num_rows = 9
        excel_data_cols = num_cols * 2
        
        # Title
        title_text = "GCE : : ERODE - 638 316 - ANNA UNIVERSITY EXAMS - \nHALL  SKETCH - NOV/DEC 2025"
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=excel_data_cols)
        title_cell = ws.cell(row=current_row, column=1, value=title_text)
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN
        current_row += 1
        
        # Hall info
        ws.cell(row=current_row, column=1, value=f"HALL NO : {hall.name}").font = HEADER_FONT
        ws.cell(row=current_row, column=4, value=f"Date & Session :  {exam_date} {session}").font = HEADER_FONT
        current_row += 1
        
        # Column headers
        for col_idx in range(num_cols):
            excel_col = (col_idx * 2) + 1
            roman = row_numerals[col_idx]
            header_text = f"{roman} - ROW with Seat No"
            cell = ws.cell(row=current_row, column=excel_col, value=header_text)
            cell.font = HEADER_FONT
            cell.alignment = WRAP_CENTER_ALIGN
            ws.column_dimensions[get_column_letter(excel_col)].width = 15
            ws.column_dimensions[get_column_letter(excel_col + 1)].width = 6
        current_row += 1
        
        # XXX row for first row (missing seats indicator)
        for col_idx in range(num_cols - 1):  # First two columns have XXX in row 0
            excel_col = (col_idx * 2) + 1
            ws.cell(row=current_row, column=excel_col, value="XXX").font = DATA_FONT
        current_row += 1
        
        # Auditorium seat numbering (special pattern)
        # Column 1: 1-8 (rows 1-8)
        # Column 2: 16-9 (rows 1-8, reversed)
        # Column 3: 17-25 (rows 0-8)
        dept_students = defaultdict(list)
        
        for row_idx in range(8):  # 8 data rows after XXX
            data_row = current_row + row_idx
            
            # Column 1: seats 1-8
            seat_num_col1 = row_idx + 1
            excel_col = 1
            
            grid_row = row_idx + 1 if row_idx + 1 < len(grid) else None
            if grid_row and grid_row < len(grid) and 0 < len(grid[grid_row]):
                seat = grid[grid_row][0]
                if seat and seat.student:
                    ws.cell(row=data_row, column=excel_col, value=seat.student.registerNumber).font = DATA_FONT
                    dept_students[seat.student.department].append(seat.student.registerNumber)
            ws.cell(row=data_row, column=excel_col + 1, value=seat_num_col1).font = DATA_FONT
            
            # Column 2: seats 16-9 (reversed)
            seat_num_col2 = 16 - row_idx
            excel_col = 3
            
            if grid_row and grid_row < len(grid) and 1 < len(grid[grid_row]):
                seat = grid[grid_row][1]
                if seat and seat.student:
                    ws.cell(row=data_row, column=excel_col, value=seat.student.registerNumber).font = DATA_FONT
                    dept_students[seat.student.department].append(seat.student.registerNumber)
            ws.cell(row=data_row, column=excel_col + 1, value=seat_num_col2).font = DATA_FONT
            
            # Column 3: seats 17-24
            seat_num_col3 = 17 + row_idx
            excel_col = 5
            
            if grid_row and grid_row < len(grid) and 2 < len(grid[grid_row]):
                seat = grid[grid_row][2]
                if seat and seat.student:
                    ws.cell(row=data_row, column=excel_col, value=seat.student.registerNumber).font = DATA_FONT
                    dept_students[seat.student.department].append(seat.student.registerNumber)
            ws.cell(row=data_row, column=excel_col + 1, value=seat_num_col3).font = DATA_FONT
        
        current_row += 8
        
        # Last row: XXX, XXX, seat 25
        ws.cell(row=current_row, column=1, value="XXX").font = DATA_FONT
        ws.cell(row=current_row, column=3, value="XXX").font = DATA_FONT
        if len(grid) > 0 and len(grid[0]) > 2:
            seat = grid[0][2]
            if seat and seat.student:
                ws.cell(row=current_row, column=5, value=seat.student.registerNumber).font = DATA_FONT
                dept_students[seat.student.department].append(seat.student.registerNumber)
        ws.cell(row=current_row, column=6, value=25).font = DATA_FONT
        current_row += 2
        
        # Department summary
        for dept, students in sorted(dept_students.items()):
            students_sorted = sorted(students)
            range_text = students_sorted[0] if len(students_sorted) == 1 else f"{students_sorted[0]} TO {students_sorted[-1]}"
            ws.cell(row=current_row, column=1, value=dept).font = DATA_FONT
            ws.cell(row=current_row, column=2, value=range_text).font = DATA_FONT
            ws.cell(row=current_row, column=6, value=len(students)).font = DATA_FONT
            current_row += 1
        
        ws.cell(row=current_row, column=5, value="TOTAL").font = HEADER_FONT
        ws.cell(row=current_row, column=6, value=hall_seating.studentsCount).font = HEADER_FONT
        current_row += 2




def generate_student_wise_excel(seating_result: SeatingResult) -> BytesIO:
    """Generate Excel file with student-wise allocation list."""
    output = BytesIO()
    data = []
    for allocation in seating_result.studentAllocation:
        data.append({
            'Registration Number': allocation.registerNumber,
            'Subject': allocation.subject,
            'Department': allocation.department,
            'Hall': allocation.hallName,
            'Seat Number': allocation.seatNumber,
            'Row': allocation.row + 1,
            'Column': allocation.col + 1
        })
    df = pd.DataFrame(data)
    df = df.sort_values(['Registration Number'])
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Student Allocations', index=False)
        worksheet = writer.sheets['Student Allocations']
        for idx, col in enumerate(df.columns):
            worksheet.column_dimensions[chr(65 + idx)].width = 20
        
        # Freeze the top row (header)
        worksheet.freeze_panes = 'A2'
        
        # Add auto-filter on the header row
        last_col_letter = chr(65 + len(df.columns) - 1)
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"
            
    output.seek(0)
    return output
